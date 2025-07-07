from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from reports.models import ReportData


class ExcleWorkbookCreaterService:
    """Сервис для создания Excel отчета."""

    def __init__(self, data: ReportData, workbook: Workbook = None, start_row: int = 1):
        self.data = data
        self.workbook = workbook or Workbook()
        self.start_row = start_row

    def _write_header(self, sheet: Worksheet) -> None:
        """Записывает заголовок в рабочий лист."""
        cell = sheet.cell(row=self.start_row, column=1, value=self.data.headers[0])
        cell.font = Font(bold=True)

        column = 2
        for item in self.data.headers[1:]:
            sheet.merge_cells(
                start_row=self.start_row,
                start_column=column,
                end_row=self.start_row,
                end_column=column + 2,
            )
            cell = sheet.cell(row=self.start_row, column=column, value=item)
            cell.font = Font(bold=True)
            column += 3
        self.start_row += 1

    def _write_rows(self, sheet: Worksheet) -> None:
        """Записывает строки отчета в рабочий лист."""
        for row in self.data.rows:
            column = 1
            cell = sheet.cell(row=self.start_row, column=column, value=row.tender_name)
            cell.font = Font(bold=True)
            column += 1

            for value in row.values:
                cell1 = sheet.cell(row=self.start_row, column=column, value=value.price)
                cell2 = sheet.cell(
                    row=self.start_row, column=column + 1, value=value.price_no_vat
                )
                cell3 = sheet.cell(
                    row=self.start_row, column=column + 2, value=value.own_currensie
                )
                if value.is_winner:
                    cell1.font = Font(bold=True)
                    cell2.font = Font(bold=True)
                    cell3.font = Font(bold=True)
                column += 3
            self.start_row += 1

    def _write_total(self, sheet: Worksheet) -> None:
        """Записывает итоговую строку в рабочий лист."""
        self.start_row += 1
        column = 1
        for row in self.data.rows:
            sheet.merge_cells(
                start_row=self.start_row,
                start_column=column,
                end_row=self.start_row,
                end_column=column + 2,
            )
            cell = sheet.cell(row=self.start_row, column=column, value=row.tender_name)
            cell.font = Font(bold=True)
            winners = [winner for winner in row.values if winner.is_winner]
            coursor = self.start_row + 1
            for item in winners:
                sheet.cell(row=coursor, column=column, value=item.price)
                sheet.cell(row=coursor, column=column + 1, value=item.price_no_vat)
                sheet.cell(row=coursor, column=column + 2, value=item.own_currensie)
                coursor += 1
            column += 3

    def workbook_creater(self, sheet_title: str = "Отчет") -> Workbook:
        """Создает Excel файл из данных отчета."""
        workbook = self.workbook
        sheet = workbook.active
        sheet.title = sheet_title
        self._write_header(sheet)
        self._write_rows(sheet)
        self._write_total(sheet)
        return workbook
